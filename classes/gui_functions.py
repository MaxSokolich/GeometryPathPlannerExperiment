from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication
import sys
from PyQt5.QtGui import QWheelEvent
from PyQt5 import QtGui
from PyQt5.QtWidgets import QWidget
from PyQt5.QtGui import QPixmap,QIcon
from PyQt5.QtCore import Qt, QTimer, PYQT_VERSION_STR
from PyQt5 import QtWidgets, QtGui, QtCore
import cv2
import os
from os.path import expanduser
import openpyxl 
import pandas as pd
from datetime import datetime
import sys
from PyQt5.QtWidgets import QApplication
import numpy as np
import cv2
import matplotlib.pyplot as plt 
import time
import platform
os.environ["SDL_JOYSTICK_ALLOW_BACKGROUND_EVENTS"] = "1"
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
try:
    import EasyPySpin
except Exception:
    pass

from classes.tracker_class import VideoThread
from classes.gui_widgets import Ui_MainWindow
from classes.arduino_class import ArduinoHandler
from classes.robot_class import Robot
from classes.cell_class import Cell
from classes.record_class import RecordThread
from classes.algorithm_class import control_algorithm
from classes.pathplanner import geo_algorithm
from classes.APG_dynamic import run_dynamic_pathplanner



class MainWindow(QtWidgets.QMainWindow):
    positionChanged = QtCore.pyqtSignal(QtCore.QPoint)

    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent=parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        
        
        #self.showMaximized()

        #resize some widgets to fit the screen better
        screen  = QtWidgets.QDesktopWidget().screenGeometry(-1)
        
        self.window_width = screen.width()
        self.window_height = screen.height()
        self.resize(self.window_width, self.window_height)
        self.display_width = self.window_width# self.ui.frameGeometry().width()

        self.displayheightratio = 0.79
        self.framesliderheightratio = 0.031
        self.textheightratio = .129
        self.tabheightratio = 0.925
        self.tabheightratio = 0.925
        
        self.aspectratio = 1041/801
        self.resize_widgets()

    
      

        self.new_dir_path = "d:\geoplanner\Tracking Data"
        if not os.path.exists(self.new_dir_path):
            os.makedirs(self.new_dir_path)

        #connect to arduino
        if "mac" in platform.platform():
            self.tbprint("Detected OS: macos")
            PORT = "/dev/cu.usbmodem11301"
           
        elif "Linux" in platform.platform():
            self.tbprint("Detected OS: Linux")
            PORT = "/dev/ttyACM0"

        elif "Windows" in platform.platform():
            self.tbprint("Detected OS:  Windows")
            PORT = "COM4"
        else:
            self.tbprint("undetected operating system")
            PORT = None
        
        self.arduino = ArduinoHandler(self.tbprint)
        self.arduino.connect(PORT)
        




        self.zoom_x, self.zoom_y, self.zoomscale, self.scrollamount = 1,0,0,0
        self.croppedresult = None
        self.currentframe = None
        self.frame_number = 0
        self.robots = []
        self.cells = []
        self.videopath = 0
        self.cap = None
        self.tracker = None
        self.recorder = None

        self.save_status = False
        self.output_workbook = None
        self.static_algorithm_status = False
        self.dynamic_algorithm_status = False
    
        
        self.total_length, self.obstacle_amount = 0,0
        
        
        
        self.setFile()

        self.ui.trackbutton.clicked.connect(self.track)

        self.ui.maskbutton.clicked.connect(self.showmask)
        self.ui.maskinvert_checkBox.toggled.connect(self.invertmaskcommand)
    
        self.ui.robotmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotcroplengthbox.valueChanged.connect(self.get_slider_vals)

        self.ui.cellmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellcroplengthbox.valueChanged.connect(self.get_slider_vals)
        


        self.ui.savedatabutton.clicked.connect(self.savedata)
        self.ui.VideoFeedLabel.installEventFilter(self)
        self.ui.recordbutton.clicked.connect(self.recordfunction_class)
        self.ui.resetdefaultbutton.clicked.connect(self.resetparams)
        self.ui.objectivebox.valueChanged.connect(self.get_objective)
        self.ui.exposurebox.valueChanged.connect(self.get_exposure)
        self.ui.croppedmasktoggle.clicked.connect(self.showcroppedoriginal)
        self.ui.croppedrecordbutton.clicked.connect(self.croppedrecordfunction)
        self.ui.choosevideobutton.clicked.connect(self.selectFile)

        self.ui.apply_static_button.clicked.connect(self.apply_static_func)
        self.ui.apply_dynamic_button.clicked.connect(self.apply_dynamic_func)
        self.ui.generatepathbutton.clicked.connect(self.generatepathfunc)

  

    def generatepathfunc(self):
        #when we click the apply algorithm button generate an instance of the geo_aglorithm and run it to output the trajectory
        image = self.tracker.robot_mask

        #sibtract robot from 
        try:
            for bot in self.tracker.robot_list:    
                x,y,w,h = bot.cropped_frame[-1]
                blank = np.zeros((w, h), dtype=np.uint8) 
                image[y:y+w , x:x+h] = blank 
        except Exception:
            pass


        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)

        
        start_point = (int(self.tracker.robot_list[-1].position_list[-1][0]), int(self.tracker.robot_list[-1].position_list[-1][1]))
        end_point = (self.tracker.robot_list[-1].trajectory[-1][0], self.tracker.robot_list[-1].trajectory[-1][1])
        print(start_point, end_point)
        

        alpha_geo = self.ui.alphabox.value() 
        safety_radius = self.ui.safetyradiusbox.value()
        deltas = [self.ui.delta_1.value(),
                  self.ui.delta_2.value(),
                  self.ui.delta_3.value(),
                  self.ui.delta_4.value(),
                  self.ui.delta_5.value(),
                  self.ui.delta_6.value(),
                  self.ui.delta_7.value(),
                  self.ui.delta_8.value(),
                  self.ui.delta_9.value(),
                  self.ui.delta_10.value()]
        
        pathplanner = geo_algorithm(image, start_point, end_point, alpha_geo, safety_radius, deltas)
        trajectory_nodes = pathplanner.run()
        print("1", trajectory_nodes)
        self.tracker.robot_list[-1].trajectory = trajectory_nodes
        print("2", self.tracker.robot_list[-1])
  

       


    def apply_static_func(self):
        if self.ui.apply_static_button.isChecked():
            self.ui.apply_static_button.setText("Stop")
            self.static_algorithm_status = True
            self.algorithm = control_algorithm()
        else:
            self.ui.apply_static_button.setText("Apply Static")
            self.static_algorithm_status = False
    


    def apply_dynamic_func(self):
        if self.ui.apply_dynamic_button.isChecked():
            self.ui.apply_dynamic_button.setText("Stop")
            self.dynamic_algorithm_status = True
            self.algorithm = control_algorithm()
        else:
            self.ui.apply_dynamic_button.setText("Apply Dynamic")
            self.dynamic_algorithm_status = False







    def update_image(self, frame, cell_mask, cell_list, robot_mask, robot_list):
        """Updates the image_label with a new opencv image"""

        #static algoruthm 
        if self.static_algorithm_status == True:
            if len(robot_list) >0: #a bot is present
                frame, Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq = self.algorithm.run_static(frame, robot_list)
                self.arduino.send(Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq)
        


        #dynamic algorithm
        elif self.dynamic_algorithm_status == True:
            
            if len(robot_list) >0: #a bot is present
                #remove the bot from the cell mask to avoid the algorithm mistaking the bot from an obstacle
         
                

                #define start and end points
                current_robot_position = (robot_list[-1].position_list[-1][0], robot_list[-1].position_list[-1][1])  #microrobots currnt position
                end = (robot_list[-1].trajectory[-1][0], robot_list[-1].trajectory[-1][1])  #target position
                
                #define algorithm parameters
                alpha_geo = self.ui.alphabox.value() 
                safety_radius = self.ui.safetyradiusbox.value()

                #define dynamic path
                frame, trajectory, self.total_length, self.obstacle_amount = run_dynamic_pathplanner(cell_mask, frame, current_robot_position, end, alpha_geo, safety_radius)

                #output actuation actions
                #if len(trajectory)==0:
                #    trajectory.append(end)

                frame, Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq = self.algorithm.run_dynamic(frame, current_robot_position, trajectory, end)
                freq = self.ui.frequency_box.value()
                self.arduino.send(Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq)

                


        else:
            self.arduino.send(0, 0, 0, 0, 0, 0, 0, 0, 0)

        #dynamic algorithm




        #DEFINE CURRENT ROBOT PARAMS TO A LIST
        if len(robot_list) > 0:
            self.robots = []
            for bot in robot_list:
                currentbot_params = [bot.frame_list[-1],
                                     bot.times[-1],
                                     bot.position_list[-1][0],bot.position_list[-1][1], 
                                     bot.velocity_list[-1][0], bot.velocity_list[-1][1],bot.velocity_list[-1][2],
                                     bot.blur_list[-1],
                                     bot.area_list[-1],
                                     bot.avg_area,
                                     bot.cropped_frame[-1][0],bot.cropped_frame[-1][1],bot.cropped_frame[-1][2],bot.cropped_frame[-1][3],
                                     self.total_length,
                                     self.obstacle_amount,
                                     bot.trajectory
                                    ]
                
                self.robots.append(currentbot_params)
        
        #IF SAVE STATUS THEN CONTINOUSLY SAVE THE CURRENT ROBOT PARAMS AND MAGNETIC FIELD PARAMS TO AN EXCEL ROWS
        if self.save_status == True:
            for (sheet, bot) in zip(self.robot_params_sheets,self.robots):
                sheet.append(bot[:-1])


        #camrea stuff
        frame = self.handle_zoom(frame)
    
        self.currentframe = frame
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
      
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(self.display_width, self.display_height, Qt.KeepAspectRatio)
        qt_img = QPixmap.fromImage(p)
       

        if self.videopath !=0:
            self.ui.frameslider.setValue(self.tracker.framenum)
        
        #also update robot info
        if len(self.robots) > 0:
            robot_diameter = round(np.sqrt(4*self.robots[-1][8]/np.pi),1)
            self.ui.vellcdnum.display(int(self.robots[-1][6]))
            self.ui.blurlcdnum.display(int(self.robots[-1][7]))
            self.ui.sizelcdnum.display(robot_diameter)
                
       
        self.ui.VideoFeedLabel.setPixmap(qt_img)

      
    


    def start_data_record(self):
        self.output_workbook = openpyxl.Workbook()
        #create sheet for robot data
        self.robot_params_sheets = []
        for i in range(len(self.robots)):
            robot_sheet = self.output_workbook.create_sheet(title= "Robot {}".format(i+1))
            robot_sheet.append(["Frame","Times","Pos X", "Pos Y", "Vel X", "Vel Y", "Vel Mag", "Blur", "Area", "Avg Area", 
                                "Cropped X","Cropped Y","Cropped W","Cropped H", "total_length", "obstacle_amount", "Path X", "Path Y"])
            self.robot_params_sheets.append(robot_sheet)
        

        #tell update_actions function to start appending data to the sheets
        self.save_status = True



    def stop_data_record(self):
        #tell update_actions function to stop appending data to the sheets
        self.save_status = False
        file_path  = os.path.join(self.new_dir_path, self.date+".xlsx")
        
        #add trajectory to file after the fact
        if self.output_workbook is not None:
            if len((self.robot_params_sheets)) > 0:
                try:
                    for i in range(len((self.robot_params_sheets))):
                        for idx,(x,y) in enumerate(self.robots[i][-1]):
                            self.robot_params_sheets[i].cell(row=idx+2, column=17).value = x
                            self.robot_params_sheets[i].cell(row=idx+2, column=18).value = y
                except Exception:
                    pass
       
            #save and close workbook
            try:
                self.output_workbook.remove(self.output_workbook["Sheet"])
                self.output_workbook.save(file_path)

                self.output_workbook.close()
                self.output_workbook = None
            except Exception:
                pass

    
    def savedata(self):
        if self.ui.savedatabutton.isChecked():
            self.ui.savedatabutton.setText("Stop")
            self.start_data_record()
        else:
            self.ui.savedatabutton.setText("Save Data")
            self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
            self.stop_data_record()
   
    def tbprint(self, text):
        #print to textbox
        self.ui.plainTextEdit.appendPlainText("$ "+ text)
    

    def convert_coords(self,pos):
        #need a way to convert the video position of mouse to the actually coordinate in the window
        newx = int(pos.x() * (self.video_width / self.display_width)) 
        newy = int(pos.y() * (self.video_height / self.display_height))
        return newx, newy

    def eventFilter(self, object, event):
        
        if object is self.ui.VideoFeedLabel: 
            if self.tracker is not None:
                
                    
                if event.type() == QtCore.QEvent.MouseButtonPress:   
                    if event.buttons() == QtCore.Qt.LeftButton:
                        newx, newy = self.convert_coords(event.pos())
                        #generate original bounding box
                        
                        #reset algorithm nodes
                   

                        if self.ui.robotmask_radio.isChecked():
                            x_1 = int(newx - self.ui.robotcroplengthbox.value()  / 2)
                            y_1 = int(newy - self.ui.robotcroplengthbox.value()  / 2)
                            w = self.ui.robotcroplengthbox.value()
                            h = self.ui.robotcroplengthbox.value()

                            robot = Robot()  # create robot instance
                            robot.add_frame(self.frame_number)
                            robot.add_time(0)
                            robot.add_position([newx,newy])
                            robot.add_velocity([0,0,0])
                            robot.add_crop([x_1, y_1, w, h])
                            robot.add_area(0)
                            robot.add_blur(0)
                        
                            robot.crop_length = self.ui.robotcroplengthbox.value()
                            self.tracker.robot_list.append(robot) #this has to include tracker.robot_list because I need to add it to that class
                        
                        elif self.ui.cellmask_radio.isChecked():
                            x_1 = int(newx - self.ui.cellcroplengthbox.value()  / 2)
                            y_1 = int(newy - self.ui.cellcroplengthbox.value()  / 2)
                            w = self.ui.cellcroplengthbox.value()
                            h = self.ui.cellcroplengthbox.value()

                            cell = Cell()  # create robot instance
                            cell.add_frame(self.frame_number)
                            cell.add_time(0)
                            cell.add_position([newx,newy])
                            cell.add_velocity([0,0,0])
                            cell.add_crop([x_1, y_1, w, h])
                            cell.add_area(0)
                            cell.add_blur(0)
                           
                            cell.crop_length = self.ui.cellcroplengthbox.value()
                            
                            self.tracker.cell_list.append(cell) #this has to include tracker.robot_list because I need to add it to that class

               
                    
                    
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = True
                        newx, newy = self.convert_coords(event.pos())
                        if len(self.tracker.robot_list) > 0:
                  
                            self.tracker.robot_list[-1].add_trajectory([newx, newy])
                
                
                    if event.buttons() == QtCore.Qt.MiddleButton: 
                        del self.tracker.robot_list[:]
                        del self.tracker.cell_list[:]
               
                        del self.robots[:]
                        del self.cells[:]
                     
                       
                    
                            
                elif event.type() == QtCore.QEvent.MouseMove:
                    self.zoom_x, self.zoom_y = self.convert_coords(event.pos())

                    if event.buttons() == QtCore.Qt.RightButton:
                        if self.drawing == True:
                            if len(self.tracker.robot_list)>0:
                                newx, newy = self.convert_coords(event.pos())
                                
                                self.tracker.robot_list[-1].add_trajectory([newx, newy])
                                
                
                elif event.type() == QtCore.QEvent.MouseButtonRelease:
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = False
                        
                if event.type() ==  QtCore.QEvent.Wheel:
                    steps = event.angleDelta().y() 
                    
                    self.scrollamount += (steps and steps / abs(steps/0.5))
                    self.scrollamount = max(min(self.scrollamount,20.0),1.0)
                    self.zoomscale = self.scrollamount

        
        return super().eventFilter(object, event)
            

    
        
        

    

    def update_croppedimage(self, frame, recorded_frame):
        """Updates the cropped image_label with a new cropped opencv image"""
        
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(310, 310, Qt.KeepAspectRatio)
        qt_cimg = QPixmap.fromImage(p)
        self.ui.CroppedVideoFeedLabel.setPixmap(qt_cimg)
        
        #recored the robots suroundings
        if self.croppedresult is not None:
            self.croppedresult.write(recorded_frame)

    

    def croppedrecordfunction(self):
        if self.cap is not None:
            if self.ui.croppedrecordbutton.isChecked():
                self.ui.croppedrecordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
                file_path  = os.path.join(self.new_dir_path, self.date+".mp4")
                self.croppedresult = cv2.VideoWriter(
                    file_path,
                    cv2.VideoWriter_fourcc(*"mp4v"),
                    int(self.videofps),    
                    (200, 200), ) 
                #start recording magnetic field and tracking data
                self.start_data_record()
            
            else:
                self.ui.croppedrecordbutton.setText("Record")
                if self.croppedresult is not None:
                    self.croppedresult.release()
                    self.croppedresult = None
                    self.tbprint("End Record, Data Saved")
                #stop and save the data when the record is over.
                self.stop_data_record()
    
         
    def recordfunction_class(self):
        if self.cap is not None:
            if self.ui.recordbutton.isChecked():
                self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
                self.recorder = RecordThread(self, self.date)
                self.recorder.recordstatus = True
                self.recorder.start()
                self.ui.recordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.start_data_record()
                
            else:
                self.recorder.stop()
                self.ui.recordbutton.setText("Record")
                self.tbprint("End Record, Data Saved")
                self.stop_data_record()


    
    def setFile(self):
        if self.videopath == 0:
            try:
                #self.cap  = cv2.VideoCapture(0) 
                self.cap = EasyPySpin.VideoCapture(0)
                self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
                self.aspectratio = (self.video_width / self.video_height)
                
                #self.cap.set(cv2.CAP_PROP_AUTO_WB, True)
                #self.cap.set(cv2.CAP_PROP_FPS, 30)
                #self.cap.set(cv2.CAP_PROP_FPS, 30)
            except Exception:
                self.cap  = cv2.VideoCapture(0) 
                self.tbprint("No EasyPySpin Camera Available")
                self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
                self.aspectratio = (self.video_width / self.video_height)
         
        else:
            self.cap  = cv2.VideoCapture(self.videopath)
            self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
            self.aspectratio = (self.video_width / self.video_height)
         

        
     
        self.tbprint("Width: {}  --  Height: {}  --  Fps: {}".format(self.video_width,self.video_height,self.videofps))
        

        self.resize_widgets()        

        if self.videopath == 0:
            self.ui.robotsizeunitslabel.setText("um")
            self.ui.robotvelocityunitslabel.setText("um/s")
        else:
            self.ui.robotsizeunitslabel.setText("px")
            self.ui.robotvelocityunitslabel.setText("px/s")
            self.totalnumframes = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.tbprint("Total Frames: {} ".format(self.totalnumframes))
            self.ui.frameslider.setGeometry(QtCore.QRect(10, self.display_height+12, self.display_width, 20))
            self.ui.frameslider.setMaximum(self.totalnumframes)
            self.ui.frameslider.show()
        
        self.ui.VideoFeedLabel.setPixmap(QtGui.QPixmap())
        

    def selectFile(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.ReadOnly
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*);;Text Files (*.txt);;Python Files (*.py)", options=options)

        if file_path:
            self.videopath = file_path
            file_info = QtCore.QFileInfo(file_path)
            file_name = file_info.fileName()
            self.ui.choosevideobutton.setText(file_name)
            self.tbprint(file_name)
        else:
            self.videopath = 0
            self.ui.choosevideobutton.setText("Live")
            self.tbprint("Using Video Camera")
        
        self.setFile()
    
    
        


    def track(self):
        if self.videopath is not None:
            if self.ui.trackbutton.isChecked():
                    
                self.setFile()
                
                self.tracker = VideoThread(self)
                self.tracker.change_pixmap_signal.connect(self.update_image)
                self.tracker.cropped_frame_signal.connect(self.update_croppedimage)
               
                self.tracker.start()

                self.ui.trackbutton.setText("Stop")
                self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
                self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
        
                
            else:
                self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
                self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
        
                if self.tracker is not None:
                    self.ui.trackbutton.setText("Track")
                    self.tracker.stop()
            

                    #reset mask button
                    self.tracker.mask_flag = False
                    self.ui.maskbutton.setText("Mask")
                    self.ui.maskbutton.setChecked(False)

              
                

                    #zero arduino commands

                    del self.tracker.robot_list[:]


            

    def showmask(self):
        if self.tracker is not None:
            if self.ui.maskbutton.isChecked():
                self.ui.maskbutton.setText("Original")
                self.tracker.mask_flag = True
            else:
                self.ui.maskbutton.setText("Mask")
                self.tracker.mask_flag = False
    
    def showcroppedoriginal(self):
        if self.tracker is not None:
            if self.ui.croppedmasktoggle.isChecked():
                self.ui.croppedmasktoggle.setText("Mask")
                self.tracker.croppedmask_flag = False
            else:
                self.ui.croppedmasktoggle.setText("Original")
                self.tracker.croppedmask_flag = True


         
    def get_objective(self):
        if self.tracker is not None:
            self.tracker.objective = self.ui.objectivebox.value()

    def get_exposure(self):
        if self.tracker is not None:
            self.tracker.exposure = self.ui.exposurebox.value()
            
    
    def invertmaskcommand(self):
        if self.tracker is not None:
            self.ui.maskinvert_checkBox.setText("Invert Mask: " + str(self.ui.maskinvert_checkBox.isChecked()))
            self.tracker.maskinvert = self.ui.maskinvert_checkBox.isChecked()

    
            

    
    
    
    

    def get_slider_vals(self):
        #alpha = self.ui.alphaspinBox.value()
        
        robotlower = self.ui.robotmasklowerbox.value() 
        robotupper = self.ui.robotmaskupperbox.value()
        robotdilation = self.ui.robotmaskdilationbox.value() 
        robotmaskblur = self.ui.robotmaskblurbox.value()
        robotcrop_length = self.ui.robotcroplengthbox.value()

        celllower = self.ui.cellmasklowerbox.value() 
        cellupper = self.ui.cellmaskupperbox.value()
        celldilation = self.ui.cellmaskdilationbox.value() 
        cellmaskblur = self.ui.cellmaskblurbox.value()
        cellcrop_length = self.ui.cellcroplengthbox.value()



        if self.tracker is not None: 

            self.tracker.robot_mask_lower = robotlower
            self.tracker.robot_mask_upper = robotupper
            self.tracker.robot_mask_dilation = robotdilation
            self.tracker.robot_mask_blur = robotmaskblur
            self.tracker.robot_crop_length = robotcrop_length

            self.tracker.cell_mask_lower = celllower
            self.tracker.cell_mask_upper = cellupper
            self.tracker.cell_mask_dilation = celldilation
            self.tracker.cell_mask_blur = cellmaskblur
            self.tracker.cell_crop_length = cellcrop_length



         
        
    def resetparams(self):
        self.ui.robotmasklowerbox.setValue(0)
        self.ui.robotmaskupperbox.setValue(128)
        self.ui.robotmaskdilationbox.setValue(0)
        self.ui.robotmaskblurbox.setValue(0)
        self.ui.robotcroplengthbox.setValue(40)
        self.ui.objectivebox.setValue(10)
        self.ui.exposurebox.setValue(5000)
        

    def resizeEvent(self, event):
        windowsize = event.size()
        self.window_width = windowsize.width()
        self.window_height = windowsize.height()
        self.resize_widgets()
 
    def resize_widgets(self):
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio

        self.display_width = int(self.display_height * self.aspectratio)

        self.ui.VideoFeedLabel.setGeometry(QtCore.QRect(10,  5,                       self.display_width,     self.display_height))
        self.ui.frameslider.setGeometry(QtCore.QRect(10,    self.display_height+12,   self.display_width,     self.framesliderheight))
        self.ui.plainTextEdit.setGeometry(QtCore.QRect(10,  self.display_height+20+self.framesliderheight,   self.display_width,     self.textheight))

        #self.ui.tabWidget.setGeometry(QtCore.QRect(12,  6,  260 ,     self.tabheight))

    def handle_zoom(self, frame):
        
        if self.zoomscale > 1:
            x = self.zoom_x
            y = self.zoom_y
            w = 300
            h = 300
            angle = 0
            
            # step 1: cropped a frame around the coord you wont to zoom into
            if y-w < 0 and x-h < 0:
                zoomedframe = frame[0:y+h , 0:x+w]
                cv2.rectangle(frame, (0, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = y
            elif x-w < 0:
                zoomedframe = frame[y-h:y+h , 0:x+w] 
                cv2.rectangle(frame, (0, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = h
            elif y-h < 0:
                zoomedframe = frame[0:y+h , x-w:x+w]
                cv2.rectangle(frame, (x-w, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = y
            else:
                zoomedframe = frame[y-h:y+h , x-w:x+w] 
                cv2.rectangle(frame, (x-w, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = h   
            
            # step 2: zoom into the zoomed frame a certain zoom amount
            rot_mat = cv2.getRotationMatrix2D((warpx,warpy), angle, self.zoomscale)
            zoomedframe = cv2.warpAffine(zoomedframe, rot_mat, zoomedframe.shape[1::-1], flags=cv2.INTER_LINEAR)

            #step 3: replace the original cropped frame with the new zoomed in cropped frame
            if y-h < 0 and x-w < 0:
                frame[0:y+h , 0:x+w] =  zoomedframe
            elif x-w < 0:
                frame[y-h:y+h , 0:x+w] =  zoomedframe
            elif y-h < 0:
                frame[0:y+h , x-w:x+w] =  zoomedframe
            else:
                frame[y-h:y+h , x-w:x+w] =  zoomedframe


        
        return frame

    def closeEvent(self, event):
        """
        called when x button is pressed
        """
        
        if self.tracker is not None:
            self.tracker.stop()
        #self.recorder.stop()

        self.arduino.close()
